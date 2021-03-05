# -*- coding: utf-8 -*-

"""
Created on 3/5/21 10:21 AM
@Author  : Justin Jiang
@Email   : jw_jiang@pku.edu.com
"""

import openpyxl
from helper import DataRemaker


def get_title_list():
    title_arr = list()
    title_arr.append("客户名称")
    for idx in range(1, 13):
        title_arr.append(str(idx) + "月")
    title_arr.append("合计")
    return title_arr


def write_to_excel(data_dict):
    # title部分
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    title_arr = get_title_list()
    for idx in range(len(title_arr)):
        sheet.cell(row=1, column=idx + 1, value=str(title_arr[idx]))

    data_res = data_remaker(data_dict)
    for customer_idx in range(len(data_res)):
        current = data_res[customer_idx]
        for current_idx in range(len(current)):
            sheet.cell(row=customer_idx + 2, column=current_idx + 1, value=current[current_idx])
    workbook.save("./result.xlsx")
    print("finished")


def data_remaker(data_dict):
    print("共有 {} 个客户的数据".format(len(data_dict)))
    customer_arr = data_dict.keys()
    result = []
    for customer in customer_arr:
        total = data_dict.get(customer)[0]  # 该客户的总数据
        month_record = data_dict.get(customer)[1]  # 各月份数据
        month_arr = month_record.keys()  # 有数据的月份列表
        format_arr = list()
        format_arr.append(customer)
        for idx in range(1, 13):
            if idx in month_arr:
                format_arr.append(month_record.get(idx))
            else:
                format_arr.append(0)
        format_arr.append(total)
        result.append(format_arr)
    return result


if __name__ == '__main__':
    data_path = './total.xlsx'
    reader = DataRemaker(data_path)
    data_dict = reader.get_all_customer_month()  # {'章先生': [6080.0, {1: 680.0, 3: 4480.0, 4: 920.0}]}
    print(data_dict)
    write_to_excel(data_dict)
